"""
Photo Verifier - Fotoğraf Doğrulama Sistemi
============================================
Saha ziyaret fotoğraflarının görüntülenmesi ve doğrulanması.
"""

from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, session
from datetime import datetime, timedelta
from functools import wraps
import os
import hashlib
import pymssql

from config import PROJECTS, get_project_config
from sources import get_source

app = Flask(__name__)
app.secret_key = 'photo-verifier-secret-key-2025'  # Production'da değiştir

def get_pv_connection():
    """PhotoVerifier veritabanı bağlantısı."""
    from config import PHOTOVERIFIER_DB
    return pymssql.connect(
        server=PHOTOVERIFIER_DB['host'],
        port=PHOTOVERIFIER_DB.get('port', 1433),
        user=PHOTOVERIFIER_DB['username'],
        password=PHOTOVERIFIER_DB['password'],
        database=PHOTOVERIFIER_DB['database']
    )

def login_required(f):
    """Login gerektiren sayfalar için decorator."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """Admin yetkisi gerektiren sayfalar için decorator."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'Admin':
            return "Yetkiniz yok", 403
        return f(*args, **kwargs)
    return decorated_function    

def get_current_user():
    """Oturum açmış kullanıcıyı getirir."""
    if 'user_id' in session:
        return {
            'id': session['user_id'],
            'username': session.get('username'),
            'display_name': session.get('display_name'),
            'role': session.get('role')
        }
    return None

def hash_password(password: str) -> str:
    """Şifreyi hashler."""
    return hashlib.sha256(password.encode()).hexdigest()

def log_event(action: str, project: str = None, details: str = None):
    """Aktivite loglar."""
    try:
        user = get_current_user()
        conn = get_pv_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO EventLogs (UserId, Username, Action, Project, Details, IpAddress)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (
            user['id'] if user else None,
            user['username'] if user else None,
            action,
            project,
            details,
            request.remote_addr
        ))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Log error: {e}")

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Kullanıcı girişi."""
    error = None
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        try:
            conn = get_pv_connection()
            cursor = conn.cursor(as_dict=True)
            cursor.execute('''
                SELECT Id, Username, PasswordHash, DisplayName, Role, IsActive
                FROM Users
                WHERE Username = %s
            ''', (username,))
            user = cursor.fetchone()
            conn.close()
            
            if user and user['IsActive']:
                if user['PasswordHash'] == hash_password(password):
                    session['user_id'] = user['Id']
                    session['username'] = user['Username']
                    session['display_name'] = user['DisplayName'] or user['Username']
                    session['role'] = user['Role']
                    
                    # Login logla
                    log_event('Login', details=f"Başarılı giriş")
                    
                    # LastLoginAt güncelle
                    conn = get_pv_connection()
                    cursor = conn.cursor()
                    cursor.execute('UPDATE Users SET LastLoginAt = GETDATE() WHERE Id = %s', (user['Id'],))
                    conn.commit()
                    conn.close()
                    
                    return redirect(url_for('index'))
                else:
                    error = 'Hatalı şifre'
                    log_event('LoginFailed', details=f"Hatalı şifre: {username}")
            elif user and not user['IsActive']:
                error = 'Hesap devre dışı'
            else:
                error = 'Kullanıcı bulunamadı'
        except Exception as e:
            error = f'Sistem hatası: {str(e)}'
            print(f"Login error: {e}")
    
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    """Kullanıcı çıkışı."""
    if 'user_id' in session:
        log_event('Logout')
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    """Ana sayfa - ilk projeye yönlendir."""
    return redirect(url_for('dashboard', project='adco'))


@app.route('/admin/users')
@admin_required
def admin_users():
    """Kullanıcı listesi."""
    conn = get_pv_connection()
    cursor = conn.cursor(as_dict=True)
    cursor.execute('''
        SELECT Id, Username, DisplayName, Email, Role, IsActive, AuthSource, CreatedAt, LastLoginAt
        FROM Users
        ORDER BY CreatedAt DESC
    ''')
    users = cursor.fetchall()
    conn.close()
    
    return render_template('admin_users.html',
                         users=users,
                         current_user=get_current_user(),
                         projects=PROJECTS)

@app.route('/admin/users/add', methods=['GET', 'POST'])
@admin_required
def admin_user_add():
    """Yeni kullanıcı ekle."""
    error = None
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        display_name = request.form.get('display_name', '').strip()
        email = request.form.get('email', '').strip()
        role = request.form.get('role', 'Viewer')
        
        if not username or not password:
            error = 'Kullanıcı adı ve şifre zorunlu'
        elif len(password) < 6:
            error = 'Şifre en az 6 karakter olmalı'
        else:
            try:
                conn = get_pv_connection()
                cursor = conn.cursor()
                
                # Kullanıcı adı var mı?
                cursor.execute('SELECT Id FROM Users WHERE Username = %s', (username,))
                if cursor.fetchone():
                    error = 'Bu kullanıcı adı zaten mevcut'
                else:
                    cursor.execute('''
                        INSERT INTO Users (Username, PasswordHash, DisplayName, Email, Role, IsActive, AuthSource)
                        VALUES (%s, %s, %s, %s, %s, 1, 'Local')
                    ''', (username, hash_password(password), display_name or username, email, role))
                    conn.commit()
                    log_event('UserCreate', details=f'Kullanıcı oluşturuldu: {username}')
                    conn.close()
                    return redirect(url_for('admin_users'))
                conn.close()
            except Exception as e:
                error = f'Sistem hatası: {str(e)}'
    
    return render_template('admin_user_form.html',
                         action='add',
                         user=None,
                         error=error,
                         current_user=get_current_user(),
                         projects=PROJECTS)

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_user_edit(user_id):
    """Kullanıcı düzenle."""
    error = None
    
    conn = get_pv_connection()
    cursor = conn.cursor(as_dict=True)
    cursor.execute('SELECT * FROM Users WHERE Id = %s', (user_id,))
    user = cursor.fetchone()
    conn.close()
    
    if not user:
        return "Kullanıcı bulunamadı", 404
    
    if request.method == 'POST':
        display_name = request.form.get('display_name', '').strip()
        email = request.form.get('email', '').strip()
        role = request.form.get('role', 'Viewer')
        is_active = request.form.get('is_active') == '1'
        new_password = request.form.get('new_password', '')
        
        try:
            conn = get_pv_connection()
            cursor = conn.cursor()
            
            if new_password:
                if len(new_password) < 6:
                    error = 'Şifre en az 6 karakter olmalı'
                else:
                    cursor.execute('''
                        UPDATE Users SET DisplayName = %s, Email = %s, Role = %s, IsActive = %s, PasswordHash = %s
                        WHERE Id = %s
                    ''', (display_name, email, role, is_active, hash_password(new_password), user_id))
            else:
                cursor.execute('''
                    UPDATE Users SET DisplayName = %s, Email = %s, Role = %s, IsActive = %s
                    WHERE Id = %s
                ''', (display_name, email, role, is_active, user_id))
            
            if not error:
                conn.commit()
                log_event('UserEdit', details=f'Kullanıcı düzenlendi: {user["Username"]}')
                conn.close()
                return redirect(url_for('admin_users'))
            conn.close()
        except Exception as e:
            error = f'Sistem hatası: {str(e)}'
    
    return render_template('admin_user_form.html',
                         action='edit',
                         user=user,
                         error=error,
                         current_user=get_current_user(),
                         projects=PROJECTS)    


@app.route('/<project>')
@login_required
def dashboard(project):
    """Proje dashboard'u."""
    if project not in PROJECTS:
        return "Proje bulunamadı", 404
    
    config = get_project_config(project)
    source = get_source(project)
    
    # Son 7 günün istatistikleri
    end_date = datetime.now().strftime('%Y-%m-%d')
    start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    
    try:
        stats = source.get_stats(start_date, end_date)
    except Exception as e:
        stats = {'error': str(e)}
    
    return render_template('dashboard.html',
                         project=project,
                         project_name=config['name'],
                         projects=PROJECTS,
                         stats=stats,
                         current_user=get_current_user())


@app.route('/<project>/photos')
@login_required
def photos(project):
    """Fotoğraf listesi - gruplandırılmış."""
    if project not in PROJECTS:
        return "Proje bulunamadı", 404
    
    config = get_project_config(project)
    source = get_source(project)
    
    # Filtreler
    photo_type = request.args.get('type', 'exhibition')
    date_from = request.args.get('from', (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d'))
    date_to = request.args.get('to', datetime.now().strftime('%Y-%m-%d'))
    user_id = request.args.get('user_id', type=int)
    customer_code = request.args.get('customer_code')
    
    # Personel ve mağaza listeleri (filtre seçenekleri için)
    try:
        personnel_list = source.get_personnel_list(date_from, date_to)
        customer_list = source.get_customer_list(date_from, date_to)
    except Exception as e:
        print(f"Liste hatası: {e}")
        personnel_list = []
        customer_list = []
    
    # Fotoğrafları getir (ziyarete göre gruplu)
    try:
        photos_grouped = source.get_photos_grouped(photo_type, date_from, date_to, user_id, customer_code)
    except Exception as e:
        photos_grouped = []
        print(f"Hata: {e}")
    
    return render_template('photos.html',
                         project=project,
                         project_name=config['name'],
                         projects=PROJECTS,
                         photo_type=photo_type,
                         photo_types=config.get('photo_tables', []),
                         photos_grouped=photos_grouped,
                         date_from=date_from,
                         date_to=date_to,
                         user_id=user_id,
                         customer_code=customer_code,
                         personnel_list=personnel_list,
                         customer_list=customer_list,
                            current_user=get_current_user())
                        
                    


@app.route('/<project>/visit/<int:visit_id>')
@login_required
def visit_detail(project, visit_id):
    """Tek bir ziyaretin tüm fotoğrafları."""
    if project not in PROJECTS:
        return "Proje bulunamadı", 404
    
    config = get_project_config(project)
    source = get_source(project)
    
    try:
        visit_data = source.get_visit_photos(visit_id)
    except Exception as e:
        visit_data = {'error': str(e)}
    
    return render_template('visit_detail.html',
                         project=project,
                         project_name=config['name'],
                         projects=PROJECTS,
                         visit=visit_data,
                            current_user=get_current_user())

@app.route('/<project>/duplicates')
@login_required
def duplicates(project):
    """Duplicate fotoğraflar sayfası."""
    if project not in PROJECTS:
        return "Proje bulunamadı", 404
    
    config = get_project_config(project)
    source = get_source(project)
    
    # Önce cache'den dene (hızlı), yoksa canlı hesapla
    if source.has_duplicate_cache():
        duplicate_groups = source.get_duplicates_from_cache()
        from_cache = True
    else:
        duplicate_groups = source.find_duplicates()
        from_cache = False
    
    return render_template('duplicates.html',
                         project=project,
                         project_name=config['name'],
                         projects=PROJECTS,
                         duplicate_groups=duplicate_groups,
                         from_cache=from_cache,
                            current_user=get_current_user())


@app.route('/<project>/reports')
@login_required
def reports(project):
    """Raporlar sayfası."""
    if project not in PROJECTS:
        return "Proje bulunamadı", 404
    
    config = get_project_config(project)
    
    return render_template('reports.html',
                         project=project,
                         project_name=config['name'],
                         projects=PROJECTS,
                            current_user=get_current_user())


# ==================== API ENDPOINTS ====================

@app.route('/api/<project>/verify', methods=['POST'])
@login_required
def api_verify(project):
    """Fotoğraf doğrulama API'si."""
    if project not in PROJECTS:
        return jsonify({'error': 'Proje bulunamadı'}), 404
    
    source = get_source(project)
    data = request.json
    
    try:
        result = source.verify_photo(
            photo_id=data['photo_id'],
            photo_type=data['photo_type'],
            status=data['status'],
            note=data.get('note', ''),
            visit_id=data.get('visit_id'),
            verified_by=session.get('user_id')
        )
        
        # Log event
        log_event(
            action='Verify',
            project=project,
            details=f"PhotoId: {data['photo_id']}, Type: {data['photo_type']}, Status: {data['status']}"
        )
        
        return jsonify({'success': True, 'result': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/<project>/stats')
@login_required
def api_stats(project):
    """İstatistik API'si."""
    if project not in PROJECTS:
        return jsonify({'error': 'Proje bulunamadı'}), 404
    
    source = get_source(project)
    
    days = int(request.args.get('days', 7))
    end_date = datetime.now().strftime('%Y-%m-%d')
    start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    
    try:
        stats = source.get_stats(start_date, end_date)
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/image/<project>/<path:image_path>')
@login_required
def serve_image(project, image_path):
    """Fotoğrafları serve et."""
    if project not in PROJECTS:
        return "Proje bulunamadı", 404
    
    config = get_project_config(project)
    base_path = config['image_path']
    
    full_path = os.path.join(base_path, image_path)
    
    if not os.path.abspath(full_path).startswith(os.path.abspath(base_path)):
        return "Erişim reddedildi", 403
    
    if os.path.exists(full_path):
        return send_file(full_path)
    else:
        return "Dosya bulunamadı", 404


# ==================== TEMPLATE FILTERS ====================

@app.template_filter('datetime')
@login_required
def format_datetime(value, format='%d.%m.%Y %H:%M'):
    if value is None:
        return ""
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value)
        except:
            return value
    return value.strftime(format)


@app.template_filter('date')
@login_required
def format_date(value, format='%d.%m.%Y'):
    if value is None:
        return ""
    if isinstance(value, str):
        try:
            value = datetime.fromisoformat(value)
        except:
            return value
    return value.strftime(format)


if __name__ == '__main__':
    print("=" * 50)
    print("📸 Photo Verifier - Fotoğraf Doğrulama Sistemi")
    print("=" * 50)
    print("http://localhost:5555")
    print("=" * 50)
    app.run(host='0.0.0.0', port=5555, debug=True)


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """Profil ve şifre değiştirme."""
    message = None
    error = None
    
    if request.method == 'POST':
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not all([current_password, new_password, confirm_password]):
            error = 'Tüm alanları doldurun'
        elif new_password != confirm_password:
            error = 'Yeni şifreler eşleşmiyor'
        elif len(new_password) < 6:
            error = 'Şifre en az 6 karakter olmalı'
        else:
            try:
                conn = get_pv_connection()
                cursor = conn.cursor(as_dict=True)
                cursor.execute('SELECT PasswordHash FROM Users WHERE Id = %s', (session['user_id'],))
                user = cursor.fetchone()
                
                if user and user['PasswordHash'] == hash_password(current_password):
                    cursor.execute('UPDATE Users SET PasswordHash = %s WHERE Id = %s', 
                                 (hash_password(new_password), session['user_id']))
                    conn.commit()
                    message = 'Şifre başarıyla güncellendi'
                    log_event('PasswordChange', details='Şifre değiştirildi')
                else:
                    error = 'Mevcut şifre hatalı'
                conn.close()
            except Exception as e:
                error = f'Sistem hatası: {str(e)}'
    
    return render_template('profile.html', 
                         message=message, 
                         error=error,
                         current_user=get_current_user(),
                         projects=PROJECTS)    

@app.route('/admin/logs')
@admin_required
def event_logs():
    """Event logları sayfası."""
    page = request.args.get('page', 1, type=int)
    per_page = 50
    offset = (page - 1) * per_page
    
    conn = get_pv_connection()
    cursor = conn.cursor(as_dict=True)
    
    # Toplam kayıt sayısı
    cursor.execute('SELECT COUNT(*) as total FROM EventLogs')
    total = cursor.fetchone()['total']
    
    # Logları getir
    cursor.execute('''
        SELECT Id, UserId, Username, Action, Project, Details, IpAddress, CreatedAt
        FROM EventLogs
        ORDER BY CreatedAt DESC
        OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
    ''', (offset, per_page))
    logs = cursor.fetchall()
    conn.close()
    
    total_pages = (total + per_page - 1) // per_page
    
    return render_template('admin_logs.html',
                         logs=logs,
                         page=page,
                         total_pages=total_pages,
                         total=total,
                         current_user=get_current_user(),
                         projects=PROJECTS)


@app.route('/<project>/reports/verifications')
@login_required
def report_verifications(project):
    """Doğrulama raporu - Excel export."""
    if project not in PROJECTS:
        return "Proje bulunamadı", 404
    
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Doğrulama verilerini al
    conn = get_pv_connection()
    cursor = conn.cursor(as_dict=True)
    cursor.execute('''
        SELECT v.PhotoId, v.PhotoType, v.VisitId, v.Status, v.Note, v.VerifiedAt,
               u.Username, u.DisplayName
        FROM Verifications v
        LEFT JOIN Users u ON v.VerifiedBy = u.Id
        WHERE v.Project = %s
        ORDER BY v.VerifiedAt DESC
    ''', (project,))
    verifications = cursor.fetchall()
    conn.close()
    
    # Proje DB'sinden fotoğraf detaylarını al
    source = get_source(project)
    
    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Doğrulama Raporu"
    
    # Header
    headers = ['Fotoğraf ID', 'Tür', 'Ziyaret ID', 'Personel', 'Mağaza Kodu', 'Mağaza Adı', 'Durum', 'Yorum', 'Doğrulayan', 'Doğrulama Tarihi']
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row_num, v in enumerate(verifications, 2):
        # Fotoğraf detaylarını al
        detail = source._get_photo_detail(v['PhotoId'], v['PhotoType'], v['VisitId'])
        
        status_text = {'approved': 'Onaylandı', 'rejected': 'Reddedildi', 'suspicious': 'Şüpheli'}.get(v['Status'], v['Status'])
        
        ws.cell(row=row_num, column=1, value=v['PhotoId'])
        ws.cell(row=row_num, column=2, value=v['PhotoType'])
        ws.cell(row=row_num, column=3, value=v['VisitId'])
        ws.cell(row=row_num, column=4, value=detail.get('personnel', ''))
        ws.cell(row=row_num, column=5, value=detail.get('customer_code', ''))
        ws.cell(row=row_num, column=6, value=detail.get('customer_name', ''))
        ws.cell(row=row_num, column=7, value=status_text)
        ws.cell(row=row_num, column=8, value=v['Note'] or '')
        ws.cell(row=row_num, column=9, value=v['DisplayName'] or v['Username'] or '')
        ws.cell(row=row_num, column=10, value=v['VerifiedAt'].strftime('%d.%m.%Y %H:%M') if v['VerifiedAt'] else '')
    
    # Kolon genişlikleri
    column_widths = [12, 12, 12, 25, 15, 35, 15, 40, 20, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{project}_dogrulama_raporu.xlsx'
    )

@app.route('/<project>/reports/duplicates')
@login_required
def report_duplicates(project):
    """Duplicate raporu - Excel export."""
    if project not in PROJECTS:
        return "Proje bulunamadı", 404
    
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import json
    
    # Duplicate verilerini al
    conn = get_pv_connection()
    cursor = conn.cursor(as_dict=True)
    cursor.execute('''
        SELECT Md5Hash, PhotoCount, Details, UpdatedAt
        FROM DuplicateCache
        WHERE Project = %s
        ORDER BY PhotoCount DESC
    ''', (project,))
    duplicates = cursor.fetchall()
    
    # Doğrulama durumlarını al
    cursor.execute('''
        SELECT PhotoId, PhotoType, Status, Note, VerifiedBy
        FROM Verifications
        WHERE Project = %s
    ''', (project,))
    verifications_list = cursor.fetchall()
    
    # Kullanıcı isimlerini al
    cursor.execute('SELECT Id, DisplayName, Username FROM Users')
    users_list = cursor.fetchall()
    conn.close()
    
    # Lookup dictionary'ler oluştur
    verifications = {(v['PhotoId'], v['PhotoType']): v for v in verifications_list}
    users = {u['Id']: u['DisplayName'] or u['Username'] for u in users_list}
    
    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicate Raporu"
    
    # Header
    headers = ['Hash', 'Tekrar', 'Fotoğraf ID', 'Tür', 'Ziyaret ID', 'Personel', 'Mağaza Kodu', 'Mağaza Adı', 'Fotoğraf Tarihi', 'Mesafe (km)', 'Doğrulama', 'Yorum', 'Doğrulayan']
    header_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row_num = 2
    for dup in duplicates:
        files = json.loads(dup['Details']) if dup['Details'] else []
        for f in files:
            photo_id = f.get('photo_id')
            photo_type = f.get('photo_type')
            
            # Doğrulama durumunu bul
            verification = verifications.get((photo_id, photo_type), {})
            status = verification.get('Status', '')
            status_text = {'approved': 'Onaylandı', 'rejected': 'Reddedildi', 'suspicious': 'Şüpheli'}.get(status, '')
            note = verification.get('Note', '')
            verified_by = users.get(verification.get('VerifiedBy'), '')
            
            ws.cell(row=row_num, column=1, value=dup['Md5Hash'][:12] + '...')
            ws.cell(row=row_num, column=2, value=dup['PhotoCount'])
            ws.cell(row=row_num, column=3, value=photo_id)
            ws.cell(row=row_num, column=4, value=photo_type)
            ws.cell(row=row_num, column=5, value=f.get('visit_id', ''))
            ws.cell(row=row_num, column=6, value=f.get('personnel', ''))
            ws.cell(row=row_num, column=7, value=f.get('customer_code', ''))
            ws.cell(row=row_num, column=8, value=f.get('customer_name', ''))
            ws.cell(row=row_num, column=9, value=str(f.get('photo_date', ''))[:19] if f.get('photo_date') else '')
            ws.cell(row=row_num, column=10, value=f.get('distance_km', ''))
            ws.cell(row=row_num, column=11, value=status_text)
            ws.cell(row=row_num, column=12, value=note)
            ws.cell(row=row_num, column=13, value=verified_by)
            row_num += 1
    
    # Kolon genişlikleri
    column_widths = [15, 8, 12, 12, 12, 25, 15, 35, 18, 12, 12, 30, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{project}_duplicate_raporu.xlsx'
    )

@app.route('/<project>/reports/distance-alerts')
@login_required
def report_distance_alerts(project):
    """Mesafe uyarı raporu - 1km+ uzaktan girişler."""
    if project not in PROJECTS:
        return "Proje bulunamadı", 404
    
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import json
    
    conn = get_pv_connection()
    cursor = conn.cursor(as_dict=True)
    cursor.execute('''
        SELECT Details
        FROM DuplicateCache
        WHERE Project = %s
    ''', (project,))
    rows = cursor.fetchall()
    conn.close()
    
    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Mesafe Uyarıları"
    
    # Header
    headers = ['Fotoğraf ID', 'Tür', 'Ziyaret ID', 'Personel', 'Mağaza Kodu', 'Mağaza Adı', 'Fotoğraf Tarihi', 'Mesafe (km)']
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row_num = 2
    for row in rows:
        files = json.loads(row['Details']) if row['Details'] else []
        for f in files:
            distance = f.get('distance_km')
            if distance and distance > 1:
                ws.cell(row=row_num, column=1, value=f.get('photo_id', ''))
                ws.cell(row=row_num, column=2, value=f.get('photo_type', ''))
                ws.cell(row=row_num, column=3, value=f.get('visit_id', ''))
                ws.cell(row=row_num, column=4, value=f.get('personnel', ''))
                ws.cell(row=row_num, column=5, value=f.get('customer_code', ''))
                ws.cell(row=row_num, column=6, value=f.get('customer_name', ''))
                ws.cell(row=row_num, column=7, value=str(f.get('photo_date', ''))[:19] if f.get('photo_date') else '')
                ws.cell(row=row_num, column=8, value=distance)
                row_num += 1
    
    # Kolon genişlikleri
    column_widths = [12, 12, 12, 25, 15, 35, 18, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{project}_mesafe_uyari_raporu.xlsx'
    )